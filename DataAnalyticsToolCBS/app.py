import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
from io import BytesIO
import openpyxl
from collections import Counter
import re
import json
import os

# Page configuration
st.set_page_config(
    page_title="Test Analytics Dashboard",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better UI
st.markdown("""
    <style>
    .main { padding: 0rem 1rem; }
    .stMetric { background-color: #f0f2f6; padding: 15px; border-radius: 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
    .metric-container { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 15px; color: white; margin-bottom: 20px; }
    h1 { color: #1e293b; font-weight: 700; }
    h2 { color: #334155; font-weight: 600; border-bottom: 2px solid #e2e8f0; padding-bottom: 10px; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { background-color: #f1f5f9; border-radius: 8px; padding: 8px 16px; font-weight: 500; }
    .stTabs [aria-selected="true"] { background-color: #4f46e5; color: white; }
    .success-box { padding: 1rem; border-radius: 0.5rem; background-color: #10b981; color: white; margin: 1rem 0; }
    .warning-box { padding: 1rem; border-radius: 0.5rem; background-color: #f59e0b; color: white; margin: 1rem 0; }
    .error-box { padding: 1rem; border-radius: 0.5rem; background-color: #ef4444; color: white; margin: 1rem 0; }
    </style>
""", unsafe_allow_html=True)

# Session state
if 'data' not in st.session_state:
    st.session_state.data = None
if 'file_uploaded' not in st.session_state:
    st.session_state.file_uploaded = False
if 'bug_tracker' not in st.session_state:
    st.session_state.bug_tracker = []
if 'next_defect_id' not in st.session_state:
    st.session_state.next_defect_id = 1001

def validate_excel_structure(df):
    required_columns = ['TC #', 'Stream', 'Domain', 'Offer ID', 'Test Scenario',
                        'Expected Result', 'Actual Result', 'Status', 'Comments',
                        'Tester Name', 'Test MSISDN', 'Test Date and Time']
    missing_columns = [c for c in required_columns if c not in df.columns]
    if missing_columns:
        return False, missing_columns
    return True, []

def load_data(file):
    try:
        df = pd.read_excel(file)
        is_valid, missing_cols = validate_excel_structure(df)
        if not is_valid:
            st.error(f"❌ Invalid file structure. Missing columns: {', '.join(missing_cols)}")
            st.info("Please upload a file with the following columns: TC #, Stream, Domain, Offer ID, Test Scenario, Expected Result, Actual Result, Status, Comments, Tester Name, Test MSISDN, Test Date and Time")
            return None
        if 'Test Date and Time' in df.columns:
            df['Test Date and Time'] = pd.to_datetime(df['Test Date and Time'], errors='coerce')
        return df
    except Exception as e:
        st.error(f"Error loading file: {str(e)}")
        return None

def create_status_pie_chart(df):
    status_counts = df['Status'].value_counts()
    colors = {'Pass': '#10b981', 'Fail': '#ef4444', 'Blocked': '#f59e0b', 'Pending': '#6366f1'}
    fig = px.pie(
        values=status_counts.values,
        names=status_counts.index,
        title='Test Status Distribution',
        color=status_counts.index,
        color_discrete_map=colors,
        hole=0.4
    )
    fig.update_traces(textposition='inside', textinfo='percent+label',
                      hovertemplate='<b>%{label}</b><br>Count: %{value}<br>Percentage: %{percent}<extra></extra>')
    fig.update_layout(showlegend=True, height=400, font=dict(size=14))
    return fig

def create_tester_bug_chart(df):
    bug_df = df[df['Status'].isin(['Fail', 'Blocked'])]
    tester_bugs = bug_df.groupby('Tester Name').size().reset_index(name='Bug Count').sort_values('Bug Count', ascending=True)
    fig = px.bar(
        tester_bugs, x='Bug Count', y='Tester Name', orientation='h',
        title='Testers by Number of Issues Found', color='Bug Count',
        color_continuous_scale='Reds', text='Bug Count'
    )
    fig.update_traces(texttemplate='%{text}', textposition='outside')
    fig.update_layout(showlegend=False, height=400, xaxis_title='Number of Issues', yaxis_title='Tester Name')
    return fig

def create_offer_status_chart(df):
    offer_status = df.groupby(['Offer ID', 'Status']).size().unstack(fill_value=0).reset_index()
    offer_status['Total'] = offer_status.sum(axis=1, numeric_only=True)
    offer_status = offer_status.nlargest(15, 'Total').drop('Total', axis=1)
    fig = go.Figure()
    colors = {'Pass': '#10b981', 'Fail': '#ef4444', 'Blocked': '#f59e0b', 'Pending': '#6366f1'}
    for status in ['Pass', 'Fail', 'Blocked', 'Pending']:
        if status in offer_status.columns:
            fig.add_trace(go.Bar(
                name=status,
                x=offer_status['Offer ID'].astype(str),
                y=offer_status[status],
                marker_color=colors.get(status, '#94a3b8'),
                text=offer_status[status],
                textposition='inside',
                texttemplate='%{text}'
            ))
    fig.update_layout(barmode='stack', title='Top 15 Offers - Test Case Results Distribution',
                      xaxis_title='Offer ID', yaxis_title='Number of Test Cases',
                      height=500, showlegend=True, hovermode='x unified')
    return fig

def display_statistics_page(df):
    st.title("📊 Testing Statistics Dashboard")
    total_tests = len(df)
    total_offers = df['Offer ID'].nunique()
    total_testers = df['Tester Name'].nunique()
    status_counts = df['Status'].value_counts()
    passed = status_counts.get('Pass', 0)
    failed = status_counts.get('Fail', 0)
    blocked = status_counts.get('Blocked', 0)
    pending = status_counts.get('Pending', 0)
    pass_rate = (passed/total_tests*100) if total_tests else 0
    fail_rate = (failed/total_tests*100) if total_tests else 0
    blocked_rate = (blocked/total_tests*100) if total_tests else 0
    pending_rate = (pending/total_tests*100) if total_tests else 0
    issues_total = failed + blocked
    issues_rate = (issues_total/total_tests*100) if total_tests else 0

    st.markdown("""
    <style>
    .metric-row { display: flex; gap: 20px; margin-bottom: 30px; }
    .metric-card { background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 20px; flex: 1; border: 1px solid #475569; box-shadow: 0 4px 6px rgba(0,0,0,0.3); }
    .metric-label { color: #94a3b8; font-size: 14px; font-weight: 500; margin-bottom: 8px; }
    .metric-value { color: #f1f5f9; font-size: 28px; font-weight: 700; margin-bottom: 8px; }
    .metric-delta { font-size: 14px; font-weight: 500; padding: 4px 8px; border-radius: 6px; display: inline-block; }
    .delta-positive { background-color: rgba(34, 197, 94, 0.2); color: #4ade80; }
    .delta-negative { background-color: rgba(239, 68, 68, 0.2); color: #f87171; }
    .delta-neutral { background-color: rgba(251, 146, 60, 0.2); color: #fb923c; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card">
            <div class="metric-label">Total Test Cases</div>
            <div class="metric-value">{total_tests:,}</div>
            <div class="metric-delta delta-positive">↑ {passed} passed ({pass_rate:.1f}%)</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">Total Offers</div>
            <div class="metric-value">{total_offers:,}</div>
            <div class="metric-delta delta-negative">↓ {failed} failed ({fail_rate:.1f}%)</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">Total Testers</div>
            <div class="metric-value">{total_testers:,}</div>
            <div class="metric-delta delta-neutral">⚠ {blocked} blocked ({blocked_rate:.1f}%)</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">Passed Tests</div>
            <div class="metric-value">{passed:,}</div>
            <div class="metric-delta delta-positive">✓ {pass_rate:.1f}%</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">Failed Tests</div>
            <div class="metric-value">{failed:,}</div>
            <div class="metric-delta delta-negative">✗ {fail_rate:.1f}%</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card"><div class="metric-label">Pass Rate</div><div class="metric-value">{pass_rate:.1f}%</div><div class="metric-delta delta-positive">↑ {passed} passed</div></div>
        <div class="metric-card"><div class="metric-label">Fail Rate</div><div class="metric-value">{fail_rate:.1f}%</div><div class="metric-delta delta-negative">↓ {failed} failed</div></div>
        <div class="metric-card"><div class="metric-label">Blocked Rate</div><div class="metric-value">{blocked_rate:.1f}%</div><div class="metric-delta delta-neutral">⚠ {blocked} blocked</div></div>
        <div class="metric-card"><div class="metric-label">Pending Rate</div><div class="metric-value">{pending_rate:.1f}%</div><div class="metric-delta delta-neutral">⏳ {pending} pending</div></div>
        <div class="metric-card"><div class="metric-label">Total Issues</div><div class="metric-value">{issues_total:,}</div><div class="metric-delta delta-negative">↓ {issues_rate:.1f}%</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1: st.plotly_chart(create_status_pie_chart(df), use_container_width=True)
    with col2: st.plotly_chart(create_tester_bug_chart(df), use_container_width=True)

    st.markdown("---")
    st.plotly_chart(create_offer_status_chart(df), use_container_width=True)

    st.markdown("---")
    st.subheader("🎯 Detailed Statistics by Category")
    tab1, tab2, tab3 = st.tabs(["By Stream", "By Domain", "By Date"])
    with tab1:
        st.dataframe(df.groupby(['Stream', 'Status']).size().unstack(fill_value=0), use_container_width=True)
    with tab2:
        st.dataframe(df.groupby(['Domain', 'Status']).size().unstack(fill_value=0), use_container_width=True)
    with tab3:
        df['Date'] = pd.to_datetime(df['Test Date and Time']).dt.date
        st.dataframe(df.groupby(['Date', 'Status']).size().unstack(fill_value=0), use_container_width=True)

def display_issues_page(df):
    st.title("🐛 Issues Tracker")
    issues_df = df[df['Status'].isin(['Fail', 'Blocked', 'Pending'])].copy()
    if issues_df.empty:
        st.success("🎉 No issues found! All tests have passed.")
        return

    issues_df['Date'] = pd.to_datetime(issues_df['Test Date and Time']).dt.date

    st.subheader("📅 Filter by Date")
    col1, col2 = st.columns(2)
    with col1:
        min_date, max_date = issues_df['Date'].min(), issues_df['Date'].max()
        selected_date = st.date_input("Select Date", value=max_date, min_value=min_date, max_value=max_date)
    with col2:
        status_filter = st.multiselect("Filter by Status", options=['Fail', 'Blocked', 'Pending'], default=['Fail', 'Blocked', 'Pending'])

    filtered_issues = issues_df[(issues_df['Date'] == selected_date) & (issues_df['Status'].isin(status_filter))]
    total_issues = len(filtered_issues)
    failed_count = (filtered_issues['Status'] == 'Fail').sum()
    blocked_count = (filtered_issues['Status'] == 'Blocked').sum()
    pending_count = (filtered_issues['Status'] == 'Pending').sum()

    st.markdown(f"""
    <style>
    .issues-metrics {{ display: flex; gap: 20px; margin: 20px 0; }}
    .issue-metric-card {{ background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 20px; flex: 1; text-align: center; border: 1px solid #475569; box-shadow: 0 4px 6px rgba(0,0,0,0.3); }}
    .issue-metric-label {{ color: #94a3b8; font-size: 14px; font-weight: 600; margin-bottom: 10px; text-transform: uppercase; }}
    .issue-metric-value {{ color: #f1f5f9; font-size: 36px; font-weight: 700; }}
    .failed-card {{ background: linear-gradient(135deg, #7f1d1d 0%, #991b1b 100%); }}
    .blocked-card {{ background: linear-gradient(135deg, #78350f 0%, #92400e 100%); }}
    .pending-card {{ background: linear-gradient(135deg, #4c1d95 0%, #5b21b6 100%); }}
    </style>
    <div class="issues-metrics">
        <div class="issue-metric-card"><div class="issue-metric-label">Total Issues</div><div class="issue-metric-value">{total_issues}</div></div>
        <div class="issue-metric-card failed-card"><div class="issue-metric-label">Failed</div><div class="issue-metric-value">{failed_count}</div></div>
        <div class="issue-metric-card blocked-card"><div class="issue-metric-label">Blocked</div><div class="issue-metric-value">{blocked_count}</div></div>
        <div class="issue-metric-card pending-card"><div class="issue-metric-label">Pending</div><div class="issue-metric-value">{pending_count}</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.subheader(f"📋 Issues for {selected_date}")
    if not filtered_issues.empty:
        display_df = filtered_issues[['Offer ID', 'Tester Name', 'Test Scenario', 'Status', 'Comments', 'Date']].copy()
        display_df['Test Scenario'] = display_df['Test Scenario'].apply(lambda x: x[:50] + '...' if len(str(x)) > 50 else x)
        display_df['Comments'] = display_df['Comments'].fillna('').apply(lambda x: x[:40] + '...' if len(str(x)) > 40 else x)
        display_df = display_df.reset_index(drop=True)
        st.dataframe(display_df, use_container_width=True, height=400)

        st.subheader("📥 Export Issues")
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            export_df = filtered_issues[['Offer ID', 'Tester Name', 'Test Scenario', 'Status', 'Comments', 'Date']].reset_index(drop=True)
            export_df.to_excel(writer, sheet_name=f'Issues_{selected_date}', index=False)
            from openpyxl.styles import Font, PatternFill, Alignment
            ws = writer.sheets[f'Issues_{selected_date}']
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        output.seek(0)
        st.download_button(
            label=f"📥 Download Issues for {selected_date} (Excel)",
            data=output,
            file_name=f"issues_{selected_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("No issues found for the selected date and filters.")

    st.markdown("---")
    st.subheader("📊 Issues Summary by Date")
    issues_by_date = issues_df.groupby(['Date', 'Status']).size().unstack(fill_value=0)
    fig = go.Figure()
    colors = {'Fail': '#ef4444', 'Blocked': '#f59e0b', 'Pending': '#6366f1'}
    for status in ['Fail', 'Blocked', 'Pending']:
        if status in issues_by_date.columns:
            fig.add_trace(go.Scatter(x=issues_by_date.index, y=issues_by_date[status],
                                     mode='lines+markers', name=status,
                                     line=dict(color=colors.get(status), width=2),
                                     marker=dict(size=8)))
    fig.update_layout(title='Issues Trend Over Time', xaxis_title='Date', yaxis_title='Number of Issues',
                      height=400, hovermode='x unified')
    st.plotly_chart(fig, use_container_width=True)

def display_comparison_page(df):
    st.title("🔄 Test Result Comparison")
    st.markdown("Identifying test cases with conflicting results from different testers")
    grouped = df.groupby(['Offer ID', 'TC #', 'Test Scenario'])
    conflicts = []
    for (offer_id, tc_num, scenario), group in grouped:
        if len(group['Tester Name'].unique()) > 1:
            statuses = group[['Tester Name', 'Status', 'Actual Result']].values
            if len(group['Status'].unique()) > 1:
                for i, row1 in enumerate(statuses):
                    for row2 in statuses[i+1:]:
                        if row1[1] != row2[1]:
                            conflicts.append({
                                'Offer ID': offer_id,
                                'TC #': tc_num,
                                'Test Scenario': (scenario[:50] + '...') if isinstance(scenario, str) and len(scenario) > 50 else scenario,
                                'Tester 1': row1[0], 'Status 1': row1[1], 'Result 1': row1[2],
                                'Tester 2': row2[0], 'Status 2': row2[1], 'Result 2': row2[2]
                            })
    if conflicts:
        conflicts_df = pd.DataFrame(conflicts)
        total_conflicts = len(conflicts_df)
        affected_offers = conflicts_df['Offer ID'].nunique()
        affected_test_cases = conflicts_df['TC #'].nunique()

        st.markdown(f"""
        <style>
        .comparison-metrics {{ display: flex; gap: 20px; margin: 20px 0 30px 0; }}
        .comparison-card {{ background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 25px; flex: 1; text-align: center; border: 1px solid #475569; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3); transition: transform 0.2s; }}
        .comparison-card:hover {{ transform: translateY(-5px); box-shadow: 0 6px 12px rgba(0, 0, 0, 0.4); }}
        .comparison-label {{ color: #94a3b8; font-size: 14px; font-weight: 600; margin-bottom: 10px; text-transform: uppercase; letter-spacing: 1px; }}
        .comparison-value {{ color: #f1f5f9; font-size: 36px; font-weight: 700; }}
        .conflicts-card {{ background: linear-gradient(135deg, #dc2626 0%, #ef4444 100%); }}
        </style>
        <div class="comparison-metrics">
            <div class="comparison-card conflicts-card"><div class="comparison-label">Total Conflicts</div><div class="comparison-value">{total_conflicts}</div></div>
            <div class="comparison-card"><div class="comparison-label">Affected Offers</div><div class="comparison-value">{affected_offers}</div></div>
            <div class="comparison-card"><div class="comparison-label">Affected Test Cases</div><div class="comparison-value">{affected_test_cases}</div></div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")
        st.subheader("🔍 Filter Conflicts")
        col1, col2 = st.columns(2)
        with col1:
            selected_offer = st.selectbox("Filter by Offer ID", options=['All'] + sorted(conflicts_df['Offer ID'].unique().tolist()), index=0)
        with col2:
            selected_tester = st.selectbox("Filter by Tester", options=['All'] + sorted(list(set(conflicts_df['Tester 1'].unique().tolist() + conflicts_df['Tester 2'].unique().tolist()))), index=0)
        filtered_conflicts = conflicts_df.copy()
        if selected_offer != 'All':
            filtered_conflicts = filtered_conflicts[filtered_conflicts['Offer ID'] == selected_offer]
        if selected_tester != 'All':
            filtered_conflicts = filtered_conflicts[(filtered_conflicts['Tester 1'] == selected_tester) | (filtered_conflicts['Tester 2'] == selected_tester)]

        st.subheader("⚠️ Conflicting Test Results")
        for idx, conflict in filtered_conflicts.iterrows():
            with st.expander(f"Conflict #{idx+1}: TC {conflict['TC #']} - Offer {conflict['Offer ID']}"):
                st.markdown(f"**Test Scenario:** {conflict['Test Scenario']}")
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**{conflict['Tester 1']}** → {conflict['Status 1']}")
                    st.write(f"Result: {conflict['Result 1']}")
                with col2:
                    st.write(f"**{conflict['Tester 2']}** → {conflict['Status 2']}")
                    st.write(f"Result: {conflict['Result 2']}")

        st.markdown("---")
        st.subheader("📈 Conflict Analysis")
        tester_conflicts = pd.concat([conflicts_df['Tester 1'].value_counts(), conflicts_df['Tester 2'].value_counts()]).groupby(level=0).sum().sort_values(ascending=False)
        fig = px.bar(x=tester_conflicts.values, y=tester_conflicts.index, orientation='h',
                     title='Testers Involved in Conflicts', labels={'x': 'Number of Conflicts', 'y': 'Tester Name'},
                     color=tester_conflicts.values, color_continuous_scale='Reds')
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.success("✅ No conflicting test results found! All testers agree on test outcomes.")

def extract_issue_patterns(comment):
    if pd.isna(comment) or comment == 'As Expected':
        return []
    comment_lower = str(comment).lower()
    patterns = []
    keywords = {
        'balance': ['balance', 'deduction', 'charge', 'tariff'],
        'api': ['api', 'interface', 'endpoint'],
        'gui': ['gui', 'display', 'screen', 'interface'],
        'network': ['network', 'connection', 'timeout', 'latency'],
        'validation': ['validation', 'verify', 'check', 'confirm'],
        'data': ['data', 'information', 'record'],
        'error': ['error', 'exception', 'fail', 'crash'],
        'performance': ['slow', 'performance', 'delay', 'lag'],
        'configuration': ['config', 'setting', 'parameter'],
        'authentication': ['auth', 'login', 'permission', 'access']
    }
    for category, words in keywords.items():
        if any(word in comment_lower for word in words):
            patterns.append(category)
    return patterns

def display_summary_page(df):
    st.title("📝 Test Summary & Analysis")
    total_tests = len(df)
    pass_rate = (df['Status'] == 'Pass').sum() / total_tests * 100 if total_tests else 0
    test_coverage = df['Offer ID'].nunique()
    team_size = df['Tester Name'].nunique()

    st.subheader("🎯 Executive Summary")
    st.markdown(f"""
    <style>
    .summary-metrics {{ display: flex; gap: 20px; margin: 20px 0 30px 0; }}
    .summary-card {{ background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 25px; flex: 1; text-align: center; border: 1px solid #475569; box-shadow: 0 4px 6px rgba(0,0,0,0.3); }}
    .summary-label {{ color: #94a3b8; font-size: 14px; font-weight: 600; margin-bottom: 10px; text-transform: uppercase; }}
    .summary-value {{ color: #f1f5f9; font-size: 32px; font-weight: 700; margin-bottom: 5px; }}
    .summary-status {{ font-size: 14px; font-weight: 600; padding: 4px 10px; border-radius: 20px; display: inline-block; margin-top: 5px; }}
    .status-good {{ background: rgba(34, 197, 94, 0.2); color: #4ade80; }}
    .status-warning {{ background: rgba(251, 146, 60, 0.2); color: #fb923c; }}
    .status-bad {{ background: rgba(239, 68, 68, 0.2); color: #f87171; }}
    </style>
    <div class="summary-metrics">
        <div class="summary-card">
            <div class="summary-label">Overall Pass Rate</div>
            <div class="summary-value">{pass_rate:.1f}%</div>
            <div class="summary-status {'status-good' if pass_rate >= 80 else 'status-warning' if pass_rate >= 60 else 'status-bad'}">
                {'✅ Excellent' if pass_rate >= 80 else '⚠️ Needs Attention' if pass_rate >= 60 else '❌ Critical'}
            </div>
        </div>
        <div class="summary-card">
            <div class="summary-label">Test Coverage</div>
            <div class="summary-value">{test_coverage}</div>
            <div class="summary-status status-good">offers tested</div>
        </div>
        <div class="summary-card">
            <div class="summary-label">Testing Team Size</div>
            <div class="summary-value">{team_size}</div>
            <div class="summary-status status-good">active testers</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    st.subheader("🐛 Issue Analysis by Offer")

    failed_tests = df[df['Status'].isin(['Fail', 'Blocked'])]
    if not failed_tests.empty:
        offer_issues = {}
        for offer_id in failed_tests['Offer ID'].unique():
            offer_data = failed_tests[failed_tests['Offer ID'] == offer_id]
            all_patterns = []
            for comment in offer_data['Actual Result']:
                all_patterns.extend(extract_issue_patterns(comment))
            pattern_counts = Counter(all_patterns)
            failed_scenarios = offer_data['Test Scenario'].value_counts().head(3)
            offer_issues[offer_id] = {
                'total_issues': len(offer_data),
                'fail_count': (offer_data['Status'] == 'Fail').sum(),
                'blocked_count': (offer_data['Status'] == 'Blocked').sum(),
                'patterns': pattern_counts,
                'top_failed_scenarios': failed_scenarios
            }

        st.markdown("""
        <style>
        .offer-analysis-card { background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 20px; margin-bottom: 20px; border: 1px solid #475569; box-shadow: 0 4px 6px rgba(0,0,0,0.3); }
        .offer-header { color: #f1f5f9; font-size: 18px; font-weight: 700; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #475569; }
        .offer-metrics { display: flex; gap: 15px; margin-bottom: 15px; }
        .offer-metric { background: rgba(30, 41, 59, 0.5); border-radius: 8px; padding: 10px 15px; flex: 1; text-align: center; }
        .offer-metric-label { color: #94a3b8; font-size: 11px; text-transform: uppercase; margin-bottom: 5px; }
        .offer-metric-value { color: #f1f5f9; font-size: 24px; font-weight: 700; }
        .section-title { color: #cbd5e1; font-size: 14px; font-weight: 600; margin: 15px 0 10px 0; }
        .issue-item { color: #e2e8f0; font-size: 13px; margin: 5px 0; padding-left: 15px; position: relative; }
        .issue-item:before { content: "•"; position: absolute; left: 0; color: #7c3aed; }
        .recommendation-item { color: #4ade80; font-size: 13px; margin: 5px 0; padding-left: 15px; position: relative; }
        .recommendation-item:before { content: "•"; position: absolute; left: 0; color: #4ade80; }
        </style>
        """, unsafe_allow_html=True)

        for offer_id, analysis in sorted(offer_issues.items(), key=lambda x: x[1]['total_issues'], reverse=True)[:10]:
            with st.expander(f"Offer {offer_id} - {analysis['total_issues']} issues"):
                st.markdown(f"""
                <div class="offer-metrics">
                    <div class="offer-metric"><div class="offer-metric-label">Total Issues</div><div class="offer-metric-value">{analysis['total_issues']}</div></div>
                    <div class="offer-metric"><div class="offer-metric-label">Failed</div><div class="offer-metric-value" style="color: #f87171;">{analysis['fail_count']}</div></div>
                    <div class="offer-metric"><div class="offer-metric-label">Blocked</div><div class="offer-metric-value" style="color: #fb923c;">{analysis['blocked_count']}</div></div>
                </div>
                """, unsafe_allow_html=True)

                if analysis['patterns']:
                    s = "<div class='section-title'>🔍 Common Issue Categories:</div>"
                    for category, count in analysis['patterns'].most_common(3):
                        s += f"<div class='issue-item'>{category.capitalize()}: {count} occurrences</div>"
                    st.markdown(s, unsafe_allow_html=True)

                if not analysis['top_failed_scenarios'].empty:
                    s = "<div class='section-title'>📋 Most Problematic Test Scenarios:</div>"
                    for scenario, count in analysis['top_failed_scenarios'].items():
                        s += f"<div class='issue-item'>{str(scenario)[:60]}... ({count} failures)</div>"
                    st.markdown(s, unsafe_allow_html=True)

                rec = "<div class='section-title'>💡 Recommendations:</div>"
                if 'balance' in analysis['patterns']: rec += "<div class='recommendation-item'>Review balance deduction logic and tariff calculations</div>"
                if 'api' in analysis['patterns']: rec += "<div class='recommendation-item'>Check API endpoints and response handling</div>"
                if 'gui' in analysis['patterns']: rec += "<div class='recommendation-item'>Verify UI components and display logic</div>"
                if 'network' in analysis['patterns']: rec += "<div class='recommendation-item'>Investigate network connectivity and timeout issues</div>"
                if 'validation' in analysis['patterns']: rec += "<div class='recommendation-item'>Review validation rules and data integrity checks</div>"
                st.markdown(rec, unsafe_allow_html=True)
    else:
        st.success("🎉 No issues found across all offers!")

    st.markdown("---")
    st.subheader("👥 Tester Performance Summary")
    tester_stats = df.groupby('Tester Name').agg({'TC #': 'count', 'Status': lambda x: (x == 'Pass').sum()}).rename(columns={'TC #': 'Total Tests', 'Status': 'Passed Tests'})
    tester_stats['Pass Rate (%)'] = (tester_stats['Passed Tests'] / tester_stats['Total Tests'] * 100).round(1)
    tester_stats = tester_stats.sort_values('Pass Rate (%)', ascending=False)

    fig = go.Figure()
    fig.add_trace(go.Bar(x=tester_stats.index, y=tester_stats['Total Tests'], name='Total Tests', marker_color='lightblue', yaxis='y', offsetgroup=1))
    fig.add_trace(go.Bar(x=tester_stats.index, y=tester_stats['Passed Tests'], name='Passed Tests', marker_color='green', yaxis='y', offsetgroup=2))
    fig.add_trace(go.Scatter(x=tester_stats.index, y=tester_stats['Pass Rate (%)'], name='Pass Rate (%)', yaxis='y2', mode='lines+markers', marker=dict(color='red', size=10), line=dict(color='red', width=2)))
    fig.update_layout(title='Tester Performance Overview', xaxis=dict(title='Tester Name'), yaxis=dict(title='Number of Tests', side='left'),
                      yaxis2=dict(title='Pass Rate (%)', overlaying='y', side='right', range=[0,100]), hovermode='x unified', barmode='group', height=500)
    st.plotly_chart(fig, use_container_width=True)
    st.dataframe(tester_stats.style.background_gradient(subset=['Pass Rate (%)'], cmap='RdYlGn'), use_container_width=True)

    st.markdown("---")
    st.subheader("📊 Test Coverage Analysis")
    col1, col2 = st.columns(2)
    with col1:
        stream_coverage = df.groupby('Stream')['Status'].value_counts().unstack(fill_value=0)
        fig_stream = px.bar(stream_coverage.T, title='Test Coverage by Stream', labels={'value': 'Count', 'index': 'Status'},
                            color_discrete_map={'Pass':'#10b981','Fail':'#ef4444','Blocked':'#f59e0b','Pending':'#6366f1'})
        st.plotly_chart(fig_stream, use_container_width=True)
    with col2:
        domain_coverage = df.groupby('Domain')['Status'].value_counts().unstack(fill_value=0)
        top_domains = domain_coverage.sum(axis=1).nlargest(10).index
        domain_coverage_top = domain_coverage.loc[top_domains]
        fig_domain = px.bar(domain_coverage_top.T, title='Test Coverage by Top 10 Domains', labels={'value': 'Count', 'index': 'Status'},
                            color_discrete_map={'Pass':'#10b981','Fail':'#ef4444','Blocked':'#f59e0b','Pending':'#6366f1'})
        st.plotly_chart(fig_domain, use_container_width=True)

def display_bug_tracker():
    st.title("🔧 Live Bug Tracker")
    st.markdown("Track and manage bugs in real-time across your team")
    st.markdown("""
    <style>
    .bug-tracker-header { background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%); border-radius: 12px; padding: 20px; margin-bottom: 20px; color: white; text-align: center; }
    .add-bug-form { background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 20px; margin-bottom: 20px; border: 1px solid #475569; }
    .bug-stats { display: flex; gap: 20px; margin-bottom: 20px; }
    .bug-stat-card { background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 20px; flex: 1; text-align: center; border: 1px solid #475569; }
    .bug-stat-label { color: #94a3b8; font-size: 14px; margin-bottom: 10px; }
    .bug-stat-value { color: #f1f5f9; font-size: 32px; font-weight: 700; }
    .status-pending { background: rgba(239, 68, 68, 0.2); color: #f87171; padding: 2px 8px; border-radius: 4px; font-weight: 600; }
    .status-resolved { background: rgba(34, 197, 94, 0.2); color: #4ade80; padding: 2px 8px; border-radius: 4px; font-weight: 600; }
    </style>
    """, unsafe_allow_html=True)

    total_bugs = len(st.session_state.bug_tracker)
    pending_bugs = len([b for b in st.session_state.bug_tracker if b['status'] == 'Pending'])
    resolved_bugs = len([b for b in st.session_state.bug_tracker if b['status'] == 'Resolved'])
    st.markdown(f"""
    <div class="bug-stats">
        <div class="bug-stat-card"><div class="bug-stat-label">TOTAL BUGS</div><div class="bug-stat-value">{total_bugs}</div></div>
        <div class="bug-stat-card" style="background: linear-gradient(135deg, #7f1d1d 0%, #991b1b 100%);"><div class="bug-stat-label">PENDING</div><div class="bug-stat-value">{pending_bugs}</div></div>
        <div class="bug-stat-card" style="background: linear-gradient(135deg, #065f46 0%, #047857 100%);"><div class="bug-stat-label">RESOLVED</div><div class="bug-stat-value">{resolved_bugs}</div></div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📝 Add New Bug", "📋 View All Bugs", "📊 Bug Analytics"])

    with tab1:
        st.subheader("Add New Bug")
        with st.form("add_bug_form", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                defect_id = st.text_input("Defect ID*", value=f"BUG-{st.session_state.next_defect_id}", placeholder="Enter unique defect ID", help="You can use the auto-generated ID or enter your own")
                offer_id = st.text_input("Offer ID*", placeholder="e.g., 40104")
            with col2:
                tested_by = st.text_input("Tested By*", placeholder="Enter tester name")
                severity = st.selectbox("Severity", ["Critical", "High", "Medium", "Low"])
            issue_description = st.text_area("Issue Description*", placeholder="Describe the bug in detail...")
            col3, col4 = st.columns(2)
            with col3:
                test_date = st.date_input("Test Date", value=datetime.now().date())
            with col4:
                environment = st.selectbox("Environment", ["Production", "Staging", "Development", "UAT"])
            submitted = st.form_submit_button("🚀 Add Bug", use_container_width=True)
            if submitted:
                if defect_id and offer_id and tested_by and issue_description:
                    existing_ids = [b['defect_id'] for b in st.session_state.bug_tracker]
                    if defect_id in existing_ids:
                        st.error(f"❌ Defect ID '{defect_id}' already exists! Please use a unique ID.")
                    else:
                        new_bug = {'defect_id': defect_id, 'offer_id': offer_id, 'issue': issue_description, 'tested_by': tested_by,
                                   'severity': severity, 'status': 'Pending', 'test_date': str(test_date), 'environment': environment,
                                   'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'resolved_at': None, 'resolution_notes': None}
                        st.session_state.bug_tracker.append(new_bug)
                        if defect_id == f"BUG-{st.session_state.next_defect_id}":
                            st.session_state.next_defect_id += 1
                        st.success(f"✅ Bug {new_bug['defect_id']} added successfully!")
                        st.balloons()
                else:
                    st.error("❌ Please fill in all required fields (Defect ID, Offer ID, Tested By, Issue Description)")

    with tab2:
        st.subheader("Bug List")
        col1, col2, col3 = st.columns(3)
        with col1: status_filter = st.selectbox("Filter by Status", ["All", "Pending", "Resolved"])
        with col2: severity_filter = st.selectbox("Filter by Severity", ["All", "Critical", "High", "Medium", "Low"])
        with col3: search_term = st.text_input("Search bugs", placeholder="Search by ID, offer, or tester...")

        filtered_bugs = st.session_state.bug_tracker.copy()
        if status_filter != "All":
            filtered_bugs = [b for b in filtered_bugs if b['status'] == status_filter]
        if severity_filter != "All":
            filtered_bugs = [b for b in filtered_bugs if b['severity'] == severity_filter]
        if search_term:
            q = search_term.lower()
            filtered_bugs = [b for b in filtered_bugs if q in b['defect_id'].lower() or q in b['offer_id'].lower() or q in b['tested_by'].lower() or q in b['issue'].lower()]

        if filtered_bugs:
            st.markdown(f"**Showing {len(filtered_bugs)} bug(s)**")
            for idx, bug in enumerate(filtered_bugs):
                original_idx = st.session_state.bug_tracker.index(bug)
                with st.expander(f"{bug['defect_id']} - Offer {bug['offer_id']} - {bug['severity']} Priority", expanded=(bug['status'] == 'Pending')):
                    col1, col2, col3 = st.columns([2, 2, 1])
                    with col1:
                        st.markdown(f"**🔍 Issue:** {bug['issue']}")
                        st.markdown(f"**👤 Tested By:** {bug['tested_by']}")
                        st.markdown(f"**📅 Test Date:** {bug['test_date']}")
                        st.markdown(f"**🌍 Environment:** {bug['environment']}")
                    with col2:
                        st.markdown(f"**⏰ Created:** {bug['created_at']}")
                        sev_emoji = {"Critical":"🔴","High":"🟠","Medium":"🟡","Low":"🟢"}.get(bug['severity'],"")
                        st.markdown(f"**Priority:** {sev_emoji} {bug['severity']}")
                        if bug['status'] == 'Resolved' and bug['resolved_at']:
                            st.markdown(f"**✅ Resolved:** {bug['resolved_at']}")
                            if bug['resolution_notes']:
                                st.markdown(f"**📝 Resolution:** {bug['resolution_notes']}")
                    with col3:
                        if bug['status'] == 'Pending':
                            st.markdown('<span class="status-pending">⚠️ PENDING</span>', unsafe_allow_html=True)
                            resolution_notes = st.text_area("Resolution notes", key=f"notes_{original_idx}", height=100)
                            if st.button("Mark Resolved", key=f"resolve_{original_idx}"):
                                st.session_state.bug_tracker[original_idx]['status'] = 'Resolved'
                                st.session_state.bug_tracker[original_idx]['resolved_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                st.session_state.bug_tracker[original_idx]['resolution_notes'] = resolution_notes or "Resolved"
                                st.success("✅ Bug marked as resolved!")
                                st.rerun()
                        else:
                            st.markdown('<span class="status-resolved">✅ RESOLVED</span>', unsafe_allow_html=True)
                            if st.button("Reopen", key=f"reopen_{original_idx}"):
                                st.session_state.bug_tracker[original_idx]['status'] = 'Pending'
                                st.session_state.bug_tracker[original_idx]['resolved_at'] = None
                                st.session_state.bug_tracker[original_idx]['resolution_notes'] = None
                                st.warning("⚠️ Bug reopened!")
                                st.rerun()

            st.markdown("---")
            c1, c2 = st.columns(2)
            with c1:
                if st.button("📥 Export Bug List to Excel", use_container_width=True):
                    df_bugs = pd.DataFrame(filtered_bugs)
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_bugs.to_excel(writer, sheet_name='Bug_Tracker', index=False)
                    output.seek(0)
                    st.download_button(
                        label="💾 Download Excel File",
                        data=output,
                        file_name=f"bug_tracker_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            with c2:
                if st.button("🗑️ Clear All Resolved Bugs", use_container_width=True):
                    st.session_state.bug_tracker = [b for b in st.session_state.bug_tracker if b['status'] != 'Resolved']
                    st.success("Resolved bugs cleared!")
                    st.rerun()
        else:
            st.info("No bugs found matching the selected filters. Add a new bug to get started!")

    with tab3:
        st.subheader("Bug Analytics")
        if st.session_state.bug_tracker:
            df_bugs = pd.DataFrame(st.session_state.bug_tracker)
            col1, col2 = st.columns(2)
            with col1:
                status_counts = df_bugs['status'].value_counts()
                fig_status = px.pie(values=status_counts.values, names=status_counts.index, title='Bugs by Status',
                                    color_discrete_map={'Pending': '#ef4444', 'Resolved': '#10b981'}, hole=0.4)
                st.plotly_chart(fig_status, use_container_width=True)
            with col2:
                severity_counts = df_bugs['severity'].value_counts()
                fig_severity = px.bar(x=severity_counts.index, y=severity_counts.values, title='Bugs by Severity',
                                      color=severity_counts.index,
                                      color_discrete_map={'Critical':'#dc2626','High':'#f97316','Medium':'#eab308','Low':'#22c55e'})
                st.plotly_chart(fig_severity, use_container_width=True)
            offer_counts = df_bugs['offer_id'].value_counts().head(10)
            if not offer_counts.empty:
                st.plotly_chart(px.bar(x=offer_counts.values, y=offer_counts.index, orientation='h',
                                       title='Top 10 Offers with Most Bugs', labels={'x':'Number of Bugs','y':'Offer ID'}), use_container_width=True)
            tester_counts = df_bugs['tested_by'].value_counts()
            if not tester_counts.empty:
                st.plotly_chart(px.bar(x=tester_counts.index, y=tester_counts.values, title='Bugs Reported by Tester',
                                       labels={'x':'Tester','y':'Number of Bugs'}), use_container_width=True)
        else:
            st.info("No bug data available for analytics. Start adding bugs to see insights!")

def display_tester_statistics(df):
    st.title("👤 Individual Tester Statistics")
    selected_tester = st.selectbox("Select a Tester", options=sorted(df['Tester Name'].unique()))
    tester_data = df[df['Tester Name'] == selected_tester]
    total_tests = len(tester_data)
    passed = (tester_data['Status'] == 'Pass').sum()
    failed = (tester_data['Status'] == 'Fail').sum()
    blocked = (tester_data['Status'] == 'Blocked').sum()
    pending = (tester_data['Status'] == 'Pending').sum()
    pass_rate = (passed/total_tests*100) if total_tests else 0
    fail_rate = (failed/total_tests*100) if total_tests else 0
    blocked_rate = (blocked/total_tests*100) if total_tests else 0
    pending_rate = (pending/total_tests*100) if total_tests else 0

    st.markdown(f"""
    <style>
    .tester-metrics {{ display: flex; gap: 20px; margin: 20px 0 30px 0; }}
    .tester-card {{ background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 20px; flex: 1; text-align: center; border: 1px solid #475569; box-shadow: 0 4px 6px rgba(0,0,0,0.3); min-width: 150px; }}
    .tester-label {{ color: #94a3b8; font-size: 13px; font-weight: 600; margin-bottom: 8px; text-transform: uppercase; }}
    .tester-value {{ color: #f1f5f9; font-size: 28px; font-weight: 700; margin-bottom: 5px; }}
    .tester-percent {{ font-size: 14px; padding: 3px 8px; border-radius: 15px; display: inline-block; font-weight: 600; }}
    .percent-pass {{ background: rgba(34,197,94,0.2); color:#4ade80; }}
    .percent-fail {{ background: rgba(239,68,68,0.2); color:#f87171; }}
    .percent-blocked {{ background: rgba(251,146,60,0.2); color:#fb923c; }}
    .percent-pending {{ background: rgba(168,85,247,0.2); color:#c084fc; }}
    .tester-name-card {{ background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%); margin-bottom: 20px; padding: 25px; border-radius: 15px; text-align: center; box-shadow: 0 6px 12px rgba(79,70,229,0.3); }}
    .tester-name {{ color: white; font-size: 24px; font-weight: 700; margin: 0; }}
    </style>
    <div class="tester-name-card"><div class="tester-name">📊 {selected_tester}'s Performance</div></div>
    <div class="tester-metrics">
        <div class="tester-card"><div class="tester-label">Total Tests</div><div class="tester-value">{total_tests}</div></div>
        <div class="tester-card"><div class="tester-label">Passed</div><div class="tester-value">{passed}</div><div class="tester-percent percent-pass">{pass_rate:.1f}%</div></div>
        <div class="tester-card"><div class="tester-label">Failed</div><div class="tester-value">{failed}</div><div class="tester-percent percent-fail">{fail_rate:.1f}%</div></div>
        <div class="tester-card"><div class="tester-label">Blocked</div><div class="tester-value">{blocked}</div><div class="tester-percent percent-blocked">{blocked_rate:.1f}%</div></div>
        <div class="tester-card"><div class="tester-label">Pending</div><div class="tester-value">{pending}</div><div class="tester-percent percent-pending">{pending_rate:.1f}%</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    tester_data['Date'] = pd.to_datetime(tester_data['Test Date and Time']).dt.date
    daily = tester_data.groupby(['Date', 'Status']).size().unstack(fill_value=0)
    fig = go.Figure()
    colors = {'Pass': '#10b981', 'Fail': '#ef4444', 'Blocked': '#f59e0b', 'Pending': '#6366f1'}
    for status in daily.columns:
        fig.add_trace(go.Scatter(x=daily.index, y=daily[status], mode='lines+markers', name=status,
                                 line=dict(color=colors.get(status), width=2), stackgroup='one'))
    fig.update_layout(title=f'{selected_tester} - Daily Test Performance', xaxis_title='Date', yaxis_title='Number of Tests', height=400, hovermode='x unified')
    st.plotly_chart(fig, use_container_width=True)
    st.subheader(f"📦 {selected_tester}'s Test Coverage by Offer")
    st.dataframe(tester_data.groupby(['Offer ID', 'Status']).size().unstack(fill_value=0), use_container_width=True)
    st.subheader("📝 Recent Test Activities")
    st.dataframe(tester_data.nlargest(10, 'Test Date and Time')[['Test Date and Time', 'Offer ID', 'Test Scenario', 'Status', 'Actual Result']], use_container_width=True)

def main():
    with st.sidebar:
        st.markdown("""
        <div style='text-align: center; padding: 10px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 10px; margin-bottom: 20px;'>
            <h2 style='color: white; margin: 0; font-size: 1.5rem;'>Data Analytics Tool for Price Plans and Offers</h2>
            <p style='color: #e0e7ff; margin: 5px 0 0 0; font-size: 0.9rem; font-style: italic;'>Sensitivity: Internal</p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("---")
        st.subheader("📁 Upload Excel File")
        uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])
        if uploaded_file is not None:
            with st.spinner('Loading data...'):
                data = load_data(uploaded_file)
                if data is not None:
                    st.session_state.data = data
                    st.session_state.file_uploaded = True
                    st.success(f"✅ Loaded {len(data)} test records")
        st.markdown("---")
        st.subheader("📍 Navigation")
        if st.session_state.file_uploaded:
            page = st.radio("Select Page", ["🏠 Home", "📊 Statistics", "🐛 Issues", "🔄 Comparison", "📝 Summary", "👤 Tester Stats", "🔧 Bug Tracker"], label_visibility="collapsed")
        else:
            page = st.radio("Select Page", ["🏠 Home", "🔧 Bug Tracker"], label_visibility="collapsed")
        st.markdown("---")
        if st.session_state.file_uploaded and st.session_state.data is not None:
            df_stats = st.session_state.data
            total_tests = len(df_stats)
            pass_percent = (df_stats['Status'] == 'Pass').sum() / total_tests * 100 if total_tests else 0
            today_tests = len(df_stats[pd.to_datetime(df_stats['Test Date and Time']).dt.date == datetime.now().date()])
            active_testers = df_stats['Tester Name'].nunique()
            st.markdown("### 📊 Live Statistics")
            st.markdown(f"**Total Tests:** {total_tests:,}  \n**Pass Rate:** {pass_percent:.1f}%  \n**Today's Tests:** {today_tests}  \n**Active Testers:** {active_testers}")
            st.markdown("---"); st.markdown("### 💡 Quick Tip")
            st.warning("Check the Comparison page to find conflicting test results!") if pass_percent < 80 else st.success("Great pass rate! Review the Summary for insights.")
        else:
            st.markdown("### 📋 Required Excel Columns")
            st.markdown("→ TC # (Test Case Number)  \n→ Offer ID  \n→ Test Scenario  \n→ Status (Pass/Fail/Blocked/Pending)  \n→ Comments  \n→ Tester Name  \n→ Test Date and Time")
            st.markdown("---"); st.markdown("### 🎨 Status Color Guide")
            col1, col2 = st.columns([1, 3])
            with col1: st.markdown("🟢"); st.markdown("🔴"); st.markdown("🟠"); st.markdown("🟣")
            with col2:
                st.markdown("**Pass** - Test Successful"); st.markdown("**Fail** - Issues Found"); st.markdown("**Blocked** - Cannot Test"); st.markdown("**Pending** - In Progress")
        st.markdown("---")
        st.caption("Built with ❤️ using Streamlit"); st.caption("Version 1.0.0")

    if page == "🏠 Home":
        st.markdown("""
        <h1 style='color: #f1f5f9; font-size: 2.5rem; font-weight: 700; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);'>
            🚀 Ready to Turn Your Test Data into Insights?
        </h1>
        """, unsafe_allow_html=True)
        st.markdown("---")
        if not st.session_state.file_uploaded:
            st.markdown("""
            <div style='background: linear-gradient(135deg, #FF6B6B 0%, #4ECDC4 100%); padding: 2.5rem; border-radius: 20px; color: white; text-align: center; box-shadow: 0 10px 30px rgba(0,0,0,0.2);'>
                <h1 style='color: white; font-size: 2.5rem; margin-bottom: 1rem;'>🎯 Let's Find Those Bugs Together!</h1>
                <p style='font-size: 1.3rem; margin-bottom: 0.5rem;'>Your testing team worked hard... Now let's see what they found! 🔍</p>
                <p style='font-size: 1.1rem; opacity: 0.95;'>Drop your Excel file and watch the magic happen ✨</p>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("---")
            st.subheader("🎮 What Can This Tool Do For You?")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown("""
                <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 1.5rem; border-radius: 15px; height: 220px; color: white; box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);'>
                    <h3 style='color: white; margin-bottom: 0.5rem;'>📊 Stats That Matter</h3>
                    <p style='font-size: 0.95rem;'>See who's the bug-hunting champion and which offers need some love!</p>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                st.markdown("""
                <div style='background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); padding: 1.5rem; border-radius: 15px; height: 220px; color: white; box-shadow: 0 5px 15px rgba(245, 87, 108, 0.4);'>
                    <h3 style='color: white; margin-bottom: 0.5rem;'>🐛 Bug Detective</h3>
                    <p style='font-size: 0.95rem;'>Track down those pesky bugs and export them before they escape!</p>
                </div>
                """, unsafe_allow_html=True)
            with col3:
                st.markdown("""
                <div style='background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); padding: 1.5rem; border-radius: 15px; height: 220px; color: white; box-shadow: 0 5px 15px rgba(79, 172, 254, 0.4);'>
                    <h3 style='color: white; margin-bottom: 0.5rem;'>🔄 Spot the Difference</h3>
                    <p style='font-size: 0.95rem;'>Find out when testers disagree - it's like a testing debate club!</p>
                </div>
                """, unsafe_allow_html=True)
            st.markdown("---")
            st.subheader("🎯 How to Get Started")
            st.markdown("""
            <div style='background-color: #1e293b; padding: 1.5rem; border-radius: 15px; color: white;'>
                <ol style='font-size: 1.05rem; line-height: 2;'>
                    <li><strong>📤 Upload Your File:</strong> Hit that upload button in the sidebar (it's waiting for you!)</li>
                    <li><strong>✅ File Check:</strong> Make sure your Excel has all the magic columns we need</li>
                    <li><strong>🎨 Pick Your View:</strong> Statistics, Issues, Comparisons - it's like Netflix for test data!</li>
                    <li><strong>💾 Export & Share:</strong> Download the juicy details and impress your team</li>
                </ol>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("---")
            st.success("💡 **Pro Tip:** Your Excel should have columns like TC #, Offer ID, Tester Name, Status, and more. We'll let you know if something's missing!")
            st.markdown("""
            <div style='text-align: center; margin-top: 2rem; padding: 1rem;'>
                <p style='color: #64748b; font-size: 1.1rem;'><i>Remember: Behind every bug found is a tester who deserves coffee ☕</i></p>
            </div>
            """, unsafe_allow_html=True)
        else:
            df = st.session_state.data
            st.success("✅ Data loaded successfully!")
            st.subheader("📊 Quick Overview")
            total_tests = len(df); total_offers = df['Offer ID'].nunique(); total_testers = df['Tester Name'].nunique()
            pass_rate = (df['Status'] == 'Pass').sum() / len(df) * 100 if len(df) else 0
            st.markdown(f"""
            <style>
            .overview-container {{ display: flex; gap: 20px; margin: 20px 0; }}
            .overview-card {{ background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 12px; padding: 25px; flex: 1; text-align: center; border: 1px solid #475569; box-shadow: 0 4px 6px rgba(0,0,0,0.3); transition: transform 0.2s; }}
            .overview-card:hover {{ transform: translateY(-5px); box-shadow: 0 6px 12px rgba(0,0,0,0.4); }}
            .overview-label {{ color: #94a3b8; font-size: 14px; font-weight: 600; margin-bottom: 10px; text-transform: uppercase; letter-spacing: 1px; }}
            .overview-value {{ color: #f1f5f9; font-size: 32px; font-weight: 700; }}
            .pass-rate {{ background: linear-gradient(135deg, #065f46 0%, #047857 100%); }}
            </style>
            <div class="overview-container">
                <div class="overview-card"><div class="overview-label">Total Tests</div><div class="overview-value">{total_tests:,}</div></div>
                <div class="overview-card"><div class="overview-label">Total Offers</div><div class="overview-value">{total_offers:,}</div></div>
                <div class="overview-card"><div class="overview-label">Total Testers</div><div class="overview-value">{total_testers:,}</div></div>
                <div class="overview-card pass-rate"><div class="overview-label">Pass Rate</div><div class="overview-value">{pass_rate:.1f}%</div></div>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("---")
            st.info("👈 Use the sidebar navigation to explore different analytics pages")

    elif page == "📊 Statistics" and st.session_state.file_uploaded:
        display_statistics_page(st.session_state.data)
    elif page == "🐛 Issues" and st.session_state.file_uploaded:
        display_issues_page(st.session_state.data)
    elif page == "🔄 Comparison" and st.session_state.file_uploaded:
        display_comparison_page(st.session_state.data)
    elif page == "📝 Summary" and st.session_state.file_uploaded:
        display_summary_page(st.session_state.data)
    elif page == "👤 Tester Stats" and st.session_state.file_uploaded:
        display_tester_statistics(st.session_state.data)
    elif page == "🔧 Bug Tracker":
        display_bug_tracker()

if __name__ == "__main__":
    main()
